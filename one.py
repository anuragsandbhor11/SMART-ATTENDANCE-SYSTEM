import cv2
import numpy as np
import os
import openpyxl  # Using openpyxl for Excel
from datetime import date



CurrentFolder = os.getcwd()  # Read current folder path
image = CurrentFolder + '\\Anurag_Sandbhor.png'  # (Placeholder for reference image

# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)

# Initialize variables
face_locations = []
face_names = []
process_this_frame = True

try:
    # Attempt to load existing workbook (if it exists)
    wb = openpyxl.load_workbook('attendence_excel.xlsx')
except FileNotFoundError:
    # Create a new workbook if not found
    wb = openpyxl.Workbook()

inp = input('Please give current subject lecture name')
if inp not in wb.sheetnames:
    sheet = wb.create_sheet(inp)
else:
    sheet = wb[inp]

# Write header row if it doesn't exist
if sheet['A1'].value is None:
    sheet['A1'] = 'Name/Date'
    sheet['B1'] = str(date.today())

row = 2  # Start writing attendance data from row 2
col = 0
already_attendence_taken = ""

# Load a reference image for illustration (not used in detection):
# person1_image = cv2.imread(image)  # Assuming image is a known face

face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')  # Load OpenCV's face detector




# Using webcam index 0 (default laptop camera)
cam = cv2.VideoCapture(0)

while True:
    # Capture a frame from the camera
    ret, frame = cam.read()

    # Check if frame is successfully captured
    if ret:
        # Display the frame
        cv2.imshow('Video', frame)













while True:
    ret, frame = video_capture.read()
    while True:
        frame = video_capture.read()
        # If frame capture fails, attempt to reinitialize the camera
    if not ret:
        print("Error: Camera not responding. Reconnecting...")
        video_capture.release()
        video_capture = cv2.VideoCapture(0)  # Retry opening camera










        

    # Convert the frame to grayscale for face detection
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Detect faces in the frame
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    # Loop through each detected face
    for (x, y, w, h) in faces:
        # Draw a rectangle around the detected face
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        # Extract the face region from the grayscale frame
        face_gray = gray[y:y + h, x:x + w]

        # You can perform face recognition here if you have known faces

        # Increment attendance for the detected face
        name = "Unknown"  # Assuming unknown face initially
        # Update attendance sheet with name and date
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=str(date.today()))
        row += 1  # Move to the next row for the next detection

    # Display the resulting frame
    cv2.imshow('Video', frame)

    # Check for key press to exit
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Save data on program exit
wb.save('attendence_excel.xlsx')

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()



